import os
import sys
import re
import openpyxl

DIR = "."

def findExcelRe(dir: str, reg: str):
    print(dir)
    for root, dirs, files in os.walk(dir):
        print(files)
        for f in files:
            if not f.endswith(".xlsx"):
                continue
            print(f)
            try:
                wb = openpyxl.load_workbook(f)

                for ws in wb:
                    for i in range(ws.min_row, ws.max_row + 1):
                        for j in range(ws.min_column, ws.max_column + 1):
                            print(i, j)

                            text = str(ws.cell(row=i, column=j).value)
                            if not text:
                                continue
                            print(text, reg)
                            rgx = re.compile(reg)
                            print(rgx.match(text))
                            if rgx.match(text) is not None:
                                print(os.path.join(root, f))
                                print(ws.title)
                                print(f"Cell: {i}, {j}")
            except IOError as e:
                print(f"cannot open {f}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        DIR = input("Folder: ")
    else:
        DIR = sys.argv[1]

    REG = r"\d{4}-\d{1,2}-\d{1,2}"
    findExcelRe(DIR, REG)

